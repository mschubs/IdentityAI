import os
import time
import random

from openpyxl import load_workbook

# from selenium import webdriver
# from selenium.webdriver.chrome.options import Options
import undetected_chromedriver as uc

import bs4


# VARIABLES
######################################################################################################################

# Remove windows-specific chromedriver.exe extension
chromedriver_path = "chromedriver"  # Path to ChromeDriver executable
xlsx_path = "data.xlsx"  # Path to Excel file with names and addresses

CURR_SCRIPT_PATH = os.path.realpath(os.path.dirname(__file__))
# Use forward slashes and os.path.join for cross-platform compatibility
profile_path = os.path.join(CURR_SCRIPT_PATH, "profile")  # Path to Chrome profile

FIRST_NAME_COL = 'A'  # (input)
LAST_NAME_COL = 'B'  # (input)
ADDRESS_COL = 'K'  # (input)
PHONEs_COLs = ['L', 'M', 'N', 'O', 'P']  # columns to output phone numbers  # (output)


######################################################################################################################


def open_chrome_with_profile():
    # Create a new Chrome session with the Chrome profile

    # options = Options()
    options = uc.ChromeOptions()
    # Use forward slashes for the profile path
    options.add_argument(f"--user-data-dir={profile_path}")

    # Create a new instance of the Chrome driver with the specified options
    # driver = webdriver.Chrome(executable_path=chromedriver_path, chrome_options=options)
    driver = uc.Chrome(options=options)
    return driver


def open_xlsx_file():
    # Open Excel file and return the workbook and worksheet

    wb = load_workbook(filename=xlsx_path)
    ws = wb.active
    return wb, ws


def write_phones_to_xlsx_file(wb, ws, phones, row):
    # Write phones to Excel file

    for i in range(len(phones)):
        ws[PHONEs_COLs[i] + str(row)].value = phones[i]

    wb.save(xlsx_path)


def extract_phones_from_page(page_source):
    # Extract phones from the page source and return them as a list of strings

    phones = []
    try:
        # find all phones
        soup = bs4.BeautifulSoup(page_source, "html.parser")
        # find all a tags with title containing "Call"
        a_tags = soup.find_all("a", title=lambda x: x and "Search people with phone number" in x)
        for a_tag in a_tags:
            phone = a_tag.text.strip()
            phones.append(phone)

        return phones

    except Exception as e:
        print(str(e))
        return phones


def main():
    driver = open_chrome_with_profile()  # Open Chrome with profile
    driver.get("https://www.fastpeoplesearch.com/")  # Navigate to FastPeopleSearch.com
    # if access denied, wait for user to enable vpn (only for the first time)
    if "Access Denied" in driver.page_source:
        print("Access Denied")
        time.sleep(60)  # Wait for the user to enable vpn extension
        driver.get("https://www.fastpeoplesearch.com/")  # Navigate to FastPeopleSearch.com
        if "Access Denied" in driver.page_source:
            return 1

    wb, ws = open_xlsx_file()  # Open the Excel file
    # for each row in the Excel file search for the person and write the phones to the Excel file
    # for row in range(2, ws.max_row + 1):
        # try searching for this person
    try:
        first_name = "Nandan"
        last_name = "Srikrishna"
        # address = ws[ADDRESS_COL + str(row)].value

        # if (first_name is None and last_name is None) or address is None:
        #     continue

        # search for this person
        first_name = first_name.replace(" ", "-")
        last_name = last_name.replace(" ", "-")
        # address = address.replace(" ", "-")
        driver.get("https://www.fastpeoplesearch.com/name/" + first_name + "-" + last_name)
        # Add random delay between 1-3 seconds to avoid detection
        time.sleep(random.uniform(1, 3))

        # Write page source to output file
        with open(f"{first_name}_{last_name}_source.html", "w", encoding="utf-8") as f:
            f.write(driver.page_source)

        # # try to get all phones for this person as a list of strings
        # phones = extract_phones_from_page(driver.page_source)
        # if phones:
        #     # write phones to Excel file
        #     print("Found " + str(len(phones)) + " phones for " + first_name + " " + last_name)
        #     write_phones_to_xlsx_file(wb, ws, phones, row)
        # else:
        #     print("No phones found for " + first_name + " " + last_name)

        # # wait 1 second before searching for the next person
        # time.sleep(1)

    except Exception as e:
        print(str(e))

    wb.close()
    driver.close()


if __name__ == "__main__":
    main()
